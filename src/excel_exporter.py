"""
Módulo para exportar datos extraídos de facturas a Excel y CSV.
Incluye funcionalidades de formateo, validación y generación de reportes.
"""

import pandas as pd
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    def __init__(self, datos: List[Dict[str, Any]], errores: List[Dict[str, Any]] = None,
                 directorio_salida: str = None, trimestre: str = "", año: str = ""):
        """
        Inicializa el exportador de Excel.

        Nueva estructura (desde v2.0):
        - Los reportes se guardan en: documentos/reportes/YYYY/XT/
        - Nombres de archivo: FACTURAS_{AÑO}_{TRIMESTRE}.xlsx

        Args:
            datos (List[Dict[str, Any]]): Lista de datos extraídos de facturas
            errores (List[Dict[str, Any]]): Lista de errores de extracción
            directorio_salida (str, optional): Directorio personalizado. Si None, usa estructura nueva.
            trimestre (str): Trimestre procesado (1T, 2T, 3T, 4T) - para organización
            año (str): Año procesado - para organización
        """
        self.datos = datos
        self.errores = errores or []
        self.trimestre = trimestre
        self.año = año
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Determinar directorio de salida
        if directorio_salida is None:
            # Nueva estructura: documentos/reportes/YYYY/XT/
            if trimestre and año:
                self.directorio_salida = os.path.join("documentos", "reportes", año, trimestre)
            else:
                # Fallback si no se especifica trimestre/año
                self.directorio_salida = os.path.join("documentos", "reportes", "sin_clasificar")
        else:
            self.directorio_salida = directorio_salida

        # Crear directorio de salida si no existe
        os.makedirs(self.directorio_salida, exist_ok=True)

    def _filtrar_columnas_estandar(self, datos: List[Dict[str, Any]], excluir_duplicados: bool = True, excluir_errores: bool = True) -> List[Dict[str, Any]]:
        """
        Filtra solo las columnas estándar (excluye las que empiezan con _).
        Opcionalmente excluye registros duplicados y con errores.

        Args:
            datos (List[Dict[str, Any]]): Datos originales con todos los campos
            excluir_duplicados (bool): Si True, excluye facturas marcadas como duplicadas
            excluir_errores (bool): Si True, excluye facturas con errores de procesamiento

        Returns:
            List[Dict[str, Any]]: Datos filtrados solo con columnas estándar
        """
        datos_filtrados = []
        for registro in datos:
            # Excluir duplicados si está activado
            if excluir_duplicados and registro.get('_Duplicado', False):
                continue

            # Excluir registros con errores si está activado
            if excluir_errores and '_Error' in registro:
                continue

            registro_filtrado = {k: v for k, v in registro.items() if not k.startswith('_')}
            datos_filtrados.append(registro_filtrado)
        return datos_filtrados

    def exportar_excel_basico(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo Excel básico usando pandas.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Filtrar solo columnas estándar (sin metadatos que empiezan con _)
        # Y excluir duplicados y errores
        datos_estandar = self._filtrar_columnas_estandar(self.datos, excluir_duplicados=True, excluir_errores=True)

        # Crear DataFrame
        df = pd.DataFrame(datos_estandar)

        # Exportar a Excel
        with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Facturas', index=False)

        print(f"OK Excel basico exportado: {ruta_completa}")
        return ruta_completa

    def exportar_excel_completo(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta TODOS los datos incluyendo metadatos (para debugging y control).

        Nueva nomenclatura: FACTURAS_DEBUG_{AÑO}_{TRIMESTRE}.xlsx

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            if self.trimestre and self.año:
                nombre_archivo = f"FACTURAS_DEBUG_{self.año}_{self.trimestre}.xlsx"
            else:
                nombre_archivo = f"facturas_completo_debug_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Crear DataFrame con TODOS los datos (sin filtrar)
        df = pd.DataFrame(self.datos)

        # Exportar a Excel
        with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Datos_Completos', index=False)

        print(f"OK Excel completo (debug) exportado: {ruta_completa}")
        return ruta_completa

    def exportar_excel_errores(self, nombre_archivo: Optional[str] = None) -> Optional[str]:
        """
        Exporta un Excel con los errores de extracción (para debugging).

        Nueva nomenclatura: ERRORES_{AÑO}_{TRIMESTRE}.xlsx

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            Optional[str]: Ruta del archivo generado o None si no hay errores
        """
        if not self.errores:
            print("No hay errores para exportar")
            return None

        if nombre_archivo is None:
            if self.trimestre and self.año:
                nombre_archivo = f"ERRORES_{self.año}_{self.trimestre}.xlsx"
            else:
                nombre_archivo = f"errores_extraccion_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Crear DataFrame con los errores
        df_errores = pd.DataFrame(self.errores)

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores"

        # Título
        ws['A1'] = "LOG DE ERRORES DE EXTRACCIÓN"
        ws['A1'].font = Font(size=16, bold=True, color="FF0000")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')

        # Información general
        ws['A3'] = "Total de errores:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = len(self.errores)

        ws['A4'] = "Fecha de generación:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Agregar datos del DataFrame
        row_start = 6
        for row_idx, row in enumerate(dataframe_to_rows(df_errores, index=False, header=True), start=row_start):
            ws.append(row)

        # Formatear encabezados
        for cell in ws[row_start]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Formatear datos
        for row in ws.iter_rows(min_row=row_start+1):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Resaltar errores en rojo claro
                cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

        # Autoajustar columnas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 10  # Ancho mínimo
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Guardar workbook
        wb.save(ruta_completa)
        print(f"OK Excel de errores (debug) exportado: {ruta_completa}")
        return ruta_completa

    def exportar_excel_formateado(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo Excel con formato profesional.

        Nueva nomenclatura: FACTURAS_{AÑO}_{TRIMESTRE}.xlsx

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            if self.trimestre and self.año:
                nombre_archivo = f"FACTURAS_{self.año}_{self.trimestre}.xlsx"
            else:
                nombre_archivo = f"facturas_formateadas_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Filtrar solo columnas estándar (sin metadatos que empiezan con _)
        # Y excluir duplicados y errores
        datos_estandar = self._filtrar_columnas_estandar(self.datos, excluir_duplicados=True, excluir_errores=True)

        # Crear DataFrame - este ya contiene solo registros exitosos sin duplicados
        df = pd.DataFrame(datos_estandar)

        # df ya está filtrado, representa solo facturas exitosas
        df_exitosos = df.copy()
        df_errores = pd.DataFrame()  # No exportamos errores en el Excel formateado

        # Crear workbook
        wb = Workbook()

        # Crear hojas
        self._crear_hoja_resumen(wb, df)
        self._crear_hoja_datos(wb, df_exitosos, "Facturas_Exitosas")

        if not df_errores.empty:
            self._crear_hoja_datos(wb, df_errores, "Facturas_Con_Errores")

        self._crear_hoja_estadisticas(wb, df)

        # Remover hoja por defecto
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Guardar workbook
        wb.save(ruta_completa)
        print(f"OK Excel formateado exportado: {ruta_completa}")
        return ruta_completa

    def _crear_hoja_resumen(self, wb: Workbook, df: pd.DataFrame):
        """Crea la hoja de resumen con información general."""
        ws = wb.active
        ws.title = "Resumen"

        # Título
        ws['A1'] = "RESUMEN DE EXTRACCIÓN DE FACTURAS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')

        # Información general
        row = 3
        # Calcular facturas exitosas, con errores y duplicadas desde datos originales
        df_original = pd.DataFrame(self.datos)

        facturas_duplicadas = len(df_original[df_original.get('_Duplicado', pd.Series([False] * len(df_original))) == True]) if '_Duplicado' in df_original.columns else 0

        if '_Error' in df_original.columns:
            facturas_exitosas = len(df_original[(df_original['_Error'].isna()) & (df_original.get('_Duplicado', False) == False)])
            facturas_con_errores = len(df_original[df_original['_Error'].notna()])
        else:
            facturas_exitosas = len(df)
            facturas_con_errores = 0

        info_general = [
            ("Fecha de procesamiento:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Total de facturas procesadas:", len(df_original)),
            ("Facturas únicas exportadas:", len(df)),
            ("Facturas exitosas:", facturas_exitosas),
            ("Facturas duplicadas (excluidas):", facturas_duplicadas),
            ("Facturas con errores:", facturas_con_errores),
        ]

        for etiqueta, valor in info_general:
            ws[f'A{row}'] = etiqueta
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = valor
            row += 1

        # Estadísticas por proveedor
        row += 2
        ws[f'A{row}'] = "ESTADÍSTICAS POR PROVEEDOR"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Cabeceras
        cabeceras = ["Proveedor ID", "Nombre Proveedor", "Facturas", "Exitosas", "Errores", "% Éxito"]
        for col, cabecera in enumerate(cabeceras, 1):
            ws.cell(row=row, column=col, value=cabecera).font = Font(bold=True)

        # Datos por proveedor
        proveedores_stats = self._calcular_estadisticas_proveedores(df)
        for stats in proveedores_stats:
            row += 1
            for col, valor in enumerate(stats.values(), 1):
                ws.cell(row=row, column=col, value=valor)

        # Autoajustar columnas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, 6):  # A hasta E
            max_length = 10  # Ancho mínimo
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _crear_hoja_datos(self, wb: Workbook, df: pd.DataFrame, nombre_hoja: str):
        """Crea una hoja con datos de facturas."""
        if df.empty:
            return

        ws = wb.create_sheet(title=nombre_hoja)

        # Agregar datos del DataFrame
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Formatear encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Formatear datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Resaltar errores en rojo
                if cell.value and str(cell.value).startswith("ERROR"):
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Autoajustar columnas
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, ws.max_column + 1):
            max_length = 10  # Ancho mínimo
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if hasattr(cell, 'value') and cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    def _crear_hoja_estadisticas(self, wb: Workbook, df: pd.DataFrame):
        """Crea una hoja con estadísticas detalladas."""
        ws = wb.create_sheet(title="Estadísticas")

        row = 1

        # Título
        ws[f'A{row}'] = "ESTADÍSTICAS DETALLADAS"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        row += 3

        # Campos más extraídos exitosamente
        ws[f'A{row}'] = "Campos con mayor tasa de éxito:"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        campos_stats = self._calcular_estadisticas_campos(df)
        for campo, stats in campos_stats.items():
            ws[f'A{row}'] = campo
            ws[f'B{row}'] = f"{stats['exitosos']}/{stats['total']} ({stats['porcentaje']}%)"
            row += 1

        row += 2

        # Archivos con errores - usar datos originales
        df_original = pd.DataFrame(self.datos)
        if '_Error' in df_original.columns:
            df_errores = df_original[df_original['_Error'].notna()]
        else:
            df_errores = pd.DataFrame()

        if not df_errores.empty:
            ws[f'A{row}'] = "Archivos con errores:"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1

            for _, factura in df_errores.iterrows():
                ws[f'A{row}'] = factura.get('_Archivo', 'N/A')
                ws[f'B{row}'] = factura.get('_Error', 'Error desconocido')
                row += 1

    def _calcular_estadisticas_proveedores(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Calcula estadísticas por proveedor."""
        stats = []
        # Usar datos originales para obtener metadatos
        df_original = pd.DataFrame(self.datos)
        if '_Proveedor_ID' not in df_original.columns:
            return stats

        proveedores = df_original['_Proveedor_ID'].value_counts()

        for proveedor_id, total in proveedores.items():
            df_proveedor = df_original[df_original['_Proveedor_ID'] == proveedor_id]

            # Contar exitosas: las que no tienen columna '_Error' o la tienen pero es NaN
            if '_Error' in df_proveedor.columns:
                exitosas = len(df_proveedor[df_proveedor['_Error'].isna()])
            else:
                exitosas = total

            errores = total - exitosas
            porcentaje = round((exitosas / total) * 100, 1) if total > 0 else 0

            nombre_proveedor = df_proveedor['_Proveedor_Nombre'].iloc[0] if '_Proveedor_Nombre' in df_proveedor.columns else 'N/A'

            stats.append({
                'proveedor_id': proveedor_id,
                'nombre_proveedor': nombre_proveedor,
                'total': total,
                'exitosas': exitosas,
                'errores': errores,
                'porcentaje': f"{porcentaje}%"
            })

        return sorted(stats, key=lambda x: x['total'], reverse=True)

    def _calcular_estadisticas_campos(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Calcula estadísticas de éxito por campo."""
        stats = {}

        # df ya está filtrado (solo columnas estándar), no necesitamos excluir nada
        campos = [col for col in df.columns]

        for campo in campos:
            if campo in df.columns:
                total = len(df)
                exitosos = len(df[(df[campo].notna()) & (~df[campo].astype(str).str.startswith('ERROR'))])
                porcentaje = round((exitosos / total) * 100, 1) if total > 0 else 0

                stats[campo] = {
                    'total': total,
                    'exitosos': exitosos,
                    'porcentaje': porcentaje
                }

        return dict(sorted(stats.items(), key=lambda x: x[1]['porcentaje'], reverse=True))

    def exportar_csv(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo CSV.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.csv"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Filtrar solo columnas estándar (sin metadatos que empiezan con _)
        # Y excluir duplicados y errores
        datos_estandar = self._filtrar_columnas_estandar(self.datos, excluir_duplicados=True, excluir_errores=True)

        # Crear DataFrame y exportar
        df = pd.DataFrame(datos_estandar)
        df.to_csv(ruta_completa, index=False, encoding='utf-8-sig', sep=';')

        print(f"OK CSV exportado: {ruta_completa}")
        return ruta_completa

    def exportar_json(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo JSON.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.json"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Filtrar solo columnas estándar (sin metadatos que empiezan con _)
        # Y excluir duplicados y errores
        datos_estandar = self._filtrar_columnas_estandar(self.datos, excluir_duplicados=True, excluir_errores=True)

        # Agregar metadatos
        exportacion = {
            "metadata": {
                "fecha_exportacion": datetime.now().isoformat(),
                "total_facturas": len(datos_estandar),
                "version": "1.0"
            },
            "facturas": datos_estandar
        }

        with open(ruta_completa, 'w', encoding='utf-8') as f:
            json.dump(exportacion, f, ensure_ascii=False, indent=2)

        print(f"OK JSON exportado: {ruta_completa}")
        return ruta_completa

    def exportar_todo(self, prefijo: Optional[str] = None) -> Dict[str, str]:
        """
        Exporta los datos en todos los formatos disponibles.

        Args:
            prefijo (str, optional): Prefijo para los nombres de archivo

        Returns:
            Dict[str, str]: Diccionario con las rutas de todos los archivos generados
        """
        resultados = {}

        prefijo = prefijo or f"facturas_{self.timestamp}"

        try:
            resultados['excel_basico'] = self.exportar_excel_basico(f"{prefijo}_basico.xlsx")
        except Exception as e:
            print(f"Error exportando Excel básico: {e}")

        try:
            resultados['excel_formateado'] = self.exportar_excel_formateado(f"{prefijo}_formateado.xlsx")
        except Exception as e:
            print(f"Error exportando Excel formateado: {e}")

        try:
            resultados['csv'] = self.exportar_csv(f"{prefijo}.csv")
        except Exception as e:
            print(f"Error exportando CSV: {e}")

        try:
            resultados['json'] = self.exportar_json(f"{prefijo}.json")
        except Exception as e:
            print(f"Error exportando JSON: {e}")

        # Exportar errores si existen
        try:
            ruta_errores = self.exportar_excel_errores(f"{prefijo}_ERRORES.xlsx")
            if ruta_errores:
                resultados['excel_errores'] = ruta_errores
        except Exception as e:
            print(f"Error exportando Excel de errores: {e}")

        return resultados


def main():
    """Función principal para testing del exportador."""
    # Datos de ejemplo para testing
    datos_ejemplo = [
        {
            "Archivo": "factura_001.pdf",
            "Proveedor_ID": "PROV_001",
            "Proveedor_Nombre": "Suministros A, S.L.",
            "NIF_Proveedor": "B12345678",
            "Fecha_Factura": "15/03/2024",
            "Numero_Factura": "FAC-2024-001",
            "Total_Factura": "1234.56",
            "Base_Imponible": "1020.30",
            "IVA": "214.26",
            "Fecha_Procesamiento": "2024-03-20 10:30:00"
        },
        {
            "Archivo": "factura_002.pdf",
            "Proveedor_ID": "PROV_001",
            "Proveedor_Nombre": "Suministros A, S.L.",
            "NIF_Proveedor": "B12345678",
            "Fecha_Factura": "ERROR",
            "Numero_Factura": "FAC-2024-002",
            "Total_Factura": "ERROR",
            "Error": "PDF corrupto",
            "Fecha_Procesamiento": "2024-03-20 10:35:00"
        }
    ]

    print("=== EXPORTADOR DE DATOS ===")
    exporter = ExcelExporter(datos_ejemplo)

    # Exportar en todos los formatos
    resultados = exporter.exportar_todo("test_facturas")

    print(f"\n=== ARCHIVOS GENERADOS ===")
    for formato, ruta in resultados.items():
        print(f"{formato}: {ruta}")


if __name__ == "__main__":
    main()