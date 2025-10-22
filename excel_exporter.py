"""
Módulo para exportar datos extraídos de facturas a Excel y CSV.
Incluye funcionalidades de formateo, validación y generación de reportes.
"""

import pandas as pd
import json
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    def __init__(self, datos: List[Dict[str, Any]], directorio_salida: str = "resultados"):
        """
        Inicializa el exportador de Excel.

        Args:
            datos (List[Dict[str, Any]]): Lista de datos extraídos de facturas
            directorio_salida (str): Directorio donde guardar los archivos generados
        """
        self.datos = datos
        self.directorio_salida = directorio_salida
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Crear directorio de salida si no existe
        os.makedirs(directorio_salida, exist_ok=True)

    def exportar_excel_basico(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo Excel básico usando pandas.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Crear DataFrame
        df = pd.DataFrame(self.datos)

        # Exportar a Excel
        with pd.ExcelWriter(ruta_completa, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Facturas', index=False)

        print(f"OK Excel basico exportado: {ruta_completa}")
        return ruta_completa

    def exportar_excel_formateado(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo Excel con formato profesional.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_formateadas_{self.timestamp}.xlsx"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Crear DataFrame
        df = pd.DataFrame(self.datos)

        # Separar datos por estado (exitosos vs errores)
        df_exitosos = df[~df.get('Error', pd.Series([False] * len(df))).notna()].copy()
        df_errores = df[df.get('Error', pd.Series([False] * len(df))).notna()].copy()

        # Crear workbook
        wb = Workbook()

        # Crear hojas
        self._crear_hoja_resumen(wb, df)
        self._crear_hoja_datos(wb, df_exitosos, "Facturas_Exitosas")

        if not df_errores.empty:
            self._crear_hoja_datos(wb, df_errores, "Facturas_Con_Errores")

        self._crear_hoja_estadisticas(wb, df)

        # Remover hoja por defecto
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Guardar workbook
        wb.save(ruta_completa)
        print(f"OK Excel formateado exportado: {ruta_completa}")
        return ruta_completa

    def _crear_hoja_resumen(self, wb: Workbook, df: pd.DataFrame):
        """Crea la hoja de resumen con información general."""
        ws = wb.active
        ws.title = "Resumen"

        # Título
        ws['A1'] = "RESUMEN DE EXTRACCIÓN DE FACTURAS"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')

        # Información general
        row = 3
        info_general = [
            ("Fecha de procesamiento:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("Total de facturas:", len(df)),
            ("Facturas exitosas:", len(df[~df.get('Error', pd.Series([False] * len(df))).notna()])),
            ("Facturas con errores:", len(df[df.get('Error', pd.Series([False] * len(df))).notna()])),
        ]

        for etiqueta, valor in info_general:
            ws[f'A{row}'] = etiqueta
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = valor
            row += 1

        # Estadísticas por proveedor
        row += 2
        ws[f'A{row}'] = "ESTADÍSTICAS POR PROVEEDOR"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        # Cabeceras
        cabeceras = ["Proveedor ID", "Nombre Proveedor", "Facturas", "Exitosas", "Errores", "% Éxito"]
        for col, cabecera in enumerate(cabeceras, 1):
            ws.cell(row=row, column=col, value=cabecera).font = Font(bold=True)

        # Datos por proveedor
        proveedores_stats = self._calcular_estadisticas_proveedores(df)
        for stats in proveedores_stats:
            row += 1
            for col, valor in enumerate(stats.values(), 1):
                ws.cell(row=row, column=col, value=valor)

        # Autoajustar columnas
        for col_idx in range(1, 6):  # A hasta E
            max_length = 10  # Ancho mínimo
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def _crear_hoja_datos(self, wb: Workbook, df: pd.DataFrame, nombre_hoja: str):
        """Crea una hoja con datos de facturas."""
        if df.empty:
            return

        ws = wb.create_sheet(title=nombre_hoja)

        # Agregar datos del DataFrame
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Formatear encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Formatear datos
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Resaltar errores en rojo
                if cell.value and str(cell.value).startswith("ERROR"):
                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        # Autoajustar columnas
        for col_idx in range(1, ws.max_column + 1):
            max_length = 10  # Ancho mínimo
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    def _crear_hoja_estadisticas(self, wb: Workbook, df: pd.DataFrame):
        """Crea una hoja con estadísticas detalladas."""
        ws = wb.create_sheet(title="Estadísticas")

        row = 1

        # Título
        ws[f'A{row}'] = "ESTADÍSTICAS DETALLADAS"
        ws[f'A{row}'].font = Font(size=16, bold=True)
        row += 3

        # Campos más extraídos exitosamente
        ws[f'A{row}'] = "Campos con mayor tasa de éxito:"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1

        campos_stats = self._calcular_estadisticas_campos(df)
        for campo, stats in campos_stats.items():
            ws[f'A{row}'] = campo
            ws[f'B{row}'] = f"{stats['exitosos']}/{stats['total']} ({stats['porcentaje']}%)"
            row += 1

        row += 2

        # Archivos con errores
        df_errores = df[df.get('Error', pd.Series([False] * len(df))).notna()]
        if not df_errores.empty:
            ws[f'A{row}'] = "Archivos con errores:"
            ws[f'A{row}'].font = Font(size=12, bold=True)
            row += 1

            for _, factura in df_errores.iterrows():
                ws[f'A{row}'] = factura.get('Archivo', 'N/A')
                ws[f'B{row}'] = factura.get('Error', 'Error desconocido')
                row += 1

    def _calcular_estadisticas_proveedores(self, df: pd.DataFrame) -> List[Dict[str, Any]]:
        """Calcula estadísticas por proveedor."""
        stats = []
        proveedores = df['Proveedor_ID'].value_counts()

        for proveedor_id, total in proveedores.items():
            df_proveedor = df[df['Proveedor_ID'] == proveedor_id]
            exitosas = len(df_proveedor[~df_proveedor.get('Error', pd.Series([False] * len(df_proveedor))).notna()])
            errores = total - exitosas
            porcentaje = round((exitosas / total) * 100, 1) if total > 0 else 0

            nombre_proveedor = df_proveedor['Proveedor_Nombre'].iloc[0] if 'Proveedor_Nombre' in df_proveedor.columns else 'N/A'

            stats.append({
                'proveedor_id': proveedor_id,
                'nombre_proveedor': nombre_proveedor,
                'total': total,
                'exitosas': exitosas,
                'errores': errores,
                'porcentaje': f"{porcentaje}%"
            })

        return sorted(stats, key=lambda x: x['total'], reverse=True)

    def _calcular_estadisticas_campos(self, df: pd.DataFrame) -> Dict[str, Dict[str, Any]]:
        """Calcula estadísticas de éxito por campo."""
        stats = {}

        # Obtener todos los campos excepto los meta
        campos_meta = {'Archivo', 'Proveedor_ID', 'Proveedor_Nombre', 'Fecha_Procesamiento', 'Error'}
        campos = [col for col in df.columns if col not in campos_meta]

        for campo in campos:
            if campo in df.columns:
                total = len(df)
                exitosos = len(df[(df[campo].notna()) & (~df[campo].astype(str).str.startswith('ERROR'))])
                porcentaje = round((exitosos / total) * 100, 1) if total > 0 else 0

                stats[campo] = {
                    'total': total,
                    'exitosos': exitosos,
                    'porcentaje': porcentaje
                }

        return dict(sorted(stats.items(), key=lambda x: x[1]['porcentaje'], reverse=True))

    def exportar_csv(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo CSV.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.csv"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Crear DataFrame y exportar
        df = pd.DataFrame(self.datos)
        df.to_csv(ruta_completa, index=False, encoding='utf-8-sig', sep=';')

        print(f"OK CSV exportado: {ruta_completa}")
        return ruta_completa

    def exportar_json(self, nombre_archivo: Optional[str] = None) -> str:
        """
        Exporta los datos a un archivo JSON.

        Args:
            nombre_archivo (str, optional): Nombre del archivo. Si None, se genera automáticamente.

        Returns:
            str: Ruta del archivo generado
        """
        if not self.datos:
            raise ValueError("No hay datos para exportar")

        if nombre_archivo is None:
            nombre_archivo = f"facturas_extraidas_{self.timestamp}.json"

        ruta_completa = os.path.join(self.directorio_salida, nombre_archivo)

        # Agregar metadatos
        exportacion = {
            "metadata": {
                "fecha_exportacion": datetime.now().isoformat(),
                "total_facturas": len(self.datos),
                "version": "1.0"
            },
            "facturas": self.datos
        }

        with open(ruta_completa, 'w', encoding='utf-8') as f:
            json.dump(exportacion, f, ensure_ascii=False, indent=2)

        print(f"OK JSON exportado: {ruta_completa}")
        return ruta_completa

    def exportar_todo(self, prefijo: Optional[str] = None) -> Dict[str, str]:
        """
        Exporta los datos en todos los formatos disponibles.

        Args:
            prefijo (str, optional): Prefijo para los nombres de archivo

        Returns:
            Dict[str, str]: Diccionario con las rutas de todos los archivos generados
        """
        resultados = {}

        prefijo = prefijo or f"facturas_{self.timestamp}"

        try:
            resultados['excel_basico'] = self.exportar_excel_basico(f"{prefijo}_basico.xlsx")
        except Exception as e:
            print(f"Error exportando Excel básico: {e}")

        try:
            resultados['excel_formateado'] = self.exportar_excel_formateado(f"{prefijo}_formateado.xlsx")
        except Exception as e:
            print(f"Error exportando Excel formateado: {e}")

        try:
            resultados['csv'] = self.exportar_csv(f"{prefijo}.csv")
        except Exception as e:
            print(f"Error exportando CSV: {e}")

        try:
            resultados['json'] = self.exportar_json(f"{prefijo}.json")
        except Exception as e:
            print(f"Error exportando JSON: {e}")

        return resultados


def main():
    """Función principal para testing del exportador."""
    # Datos de ejemplo para testing
    datos_ejemplo = [
        {
            "Archivo": "factura_001.pdf",
            "Proveedor_ID": "PROV_001",
            "Proveedor_Nombre": "Suministros A, S.L.",
            "NIF_Proveedor": "B12345678",
            "Fecha_Factura": "15/03/2024",
            "Numero_Factura": "FAC-2024-001",
            "Total_Factura": "1234.56",
            "Base_Imponible": "1020.30",
            "IVA": "214.26",
            "Fecha_Procesamiento": "2024-03-20 10:30:00"
        },
        {
            "Archivo": "factura_002.pdf",
            "Proveedor_ID": "PROV_001",
            "Proveedor_Nombre": "Suministros A, S.L.",
            "NIF_Proveedor": "B12345678",
            "Fecha_Factura": "ERROR",
            "Numero_Factura": "FAC-2024-002",
            "Total_Factura": "ERROR",
            "Error": "PDF corrupto",
            "Fecha_Procesamiento": "2024-03-20 10:35:00"
        }
    ]

    print("=== EXPORTADOR DE DATOS ===")
    exporter = ExcelExporter(datos_ejemplo)

    # Exportar en todos los formatos
    resultados = exporter.exportar_todo("test_facturas")

    print(f"\n=== ARCHIVOS GENERADOS ===")
    for formato, ruta in resultados.items():
        print(f"{formato}: {ruta}")


if __name__ == "__main__":
    main()